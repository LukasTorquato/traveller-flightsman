import json
import subprocess
import sys
from pathlib import Path

import openpyxl


def test_jsonl_to_xlsx_round_trip(tmp_path: Path):
    src = tmp_path / "observations.jsonl"
    dst = tmp_path / "observations.xlsx"
    rows = [
        {
            "run_date": "2026-04-21",
            "origin": "DUB",
            "destination_iata": "BCN",
            "destination_city": "Barcelona",
            "departure_date": "2026-06-12",
            "return_date": "2026-06-15",
            "nights": 3,
            "price_eur": 48.5,
            "airline": "FR",
            "stops": 0,
            "source": "kiwi",
            "is_wishlist": False,
            "category": "europe_short_haul",
            "market_p15_eur": 62.0,
            "was_flagged_as_deal": True,
            "flag_reason": "x",
            "baseline_median_eur": None,
            "phase": 1,
        },
        {
            "kind": "run_metadata",
            "run_date": "2026-04-21",
            "run_started_at": "x",
            "run_ended_at": "y",
            "total_routes_queried": 1,
            "total_api_calls": 1,
            "deals_flagged": 1,
            "errors": [],
            "git_commit_sha": None,
        },
    ]
    src.write_text("\n".join(json.dumps(r) for r in rows), encoding="utf-8")

    r = subprocess.run(
        [sys.executable, "scripts/jsonl_to_xlsx.py", str(src), str(dst)],
        capture_output=True,
        text=True,
    )
    assert r.returncode == 0, r.stderr
    wb = openpyxl.load_workbook(dst)
    assert "observations" in wb.sheetnames
    assert "run_metadata" in wb.sheetnames
    obs_ws = wb["observations"]
    assert obs_ws.cell(row=1, column=1).value == "run_date"
    assert obs_ws.cell(row=2, column=3).value == "BCN"
